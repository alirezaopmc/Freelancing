from Grid.network import Network
from openpyxl import Workbook

m = int(input("Enter the number of rows "))
n = int(input("Enter the number of columns "))

a = int(input("Enter the horizontal length "))
b = int(input("Enter the vertical length "))

path = input("Enter the output excel file name ")

lower = input("Enter the lower bound (press enter to use defaults 0) ")
upper = input("Enter the upper bound (press enter to use defaults 1000) ")

# m, n, a, b, path, lower, upper = 2, 3, 4, 5, 'hell.xlsx', '0', '1000'
net = Network(m, n, a, b, path, lower, upper)

wb = Workbook(write_only=True)

travel_times = wb.create_sheet()
travel_times.title = f'{m}x{n} Travel Times'
demand = wb.create_sheet()
demand.title = f'{m}x{n} Demand'

travel_times.append(['From', 'To', 'Length (km)', 'Line'])

row1 = 2

for value in range(m * n):
    i = value // n
    j = value % n

    adj_list = sorted(net.vertices[(i, j)].adj.items())

    for adj in adj_list:
        to_value = adj[0]
        edge = net.get_edge(value, to_value)
        length = edge.length
        line = edge.line
        travel_times.append([value+1, to_value+1, length, line])
        row1 += 1


demand.append(['From', 'To', 'Deman'])

row2 = 2
for value1 in range(m * n):
    for value2 in range(n * m):
        demand.append([value1+1, value2+1, net.get_demand(value1, value2)])
        row2 += 1


wb.save(path)
