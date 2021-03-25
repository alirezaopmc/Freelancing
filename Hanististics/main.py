from Grid.network import Network
from openpyxl import Workbook

m = int(input("Enter the number of rows "))
n = int(input("Enter the number of columns "))

a = int(input("Enter the horizontal length "))
b = int(input("Enter the vertical length "))

path = input("Enter the output excel file name ")

lower = input("Enter the lower bound (press enter to use defaults 0) ")
upper = input("Enter the upper bound (press enter to use defaults 1000) ")

# m, n, a, b, path, lower, upper = 2, 3, 4, 5, 'hell.xlsx', '0', '1000'
net = Network(m, n, a, b, path, lower, upper)

wb = Workbook()

travel_times = wb.active
travel_times.title = f'{m}x{n} Travel Times'
demand = wb.create_sheet()
demand.title = f'{m}x{n} Demand'

travel_times.cell(1, 1, 'From')
travel_times.cell(1, 2, 'To')
travel_times.cell(1, 3, 'Length (km)')
travel_times.cell(1, 4, 'Line')

row1 = 2

for value in range(m * n):
    i = value // n
    j = value % n

    adj_list = sorted(net.vertices[(i, j)].adj.items())

    for adj in adj_list:
        to_value = adj[0]
        edge = net.get_edge(value, to_value)
        length = edge.length
        line = edge.line
        travel_times.cell(row1, 1, value+1)
        travel_times.cell(row1, 2, to_value+1)
        travel_times.cell(row1, 3, length)
        travel_times.cell(row1, 4, line)
        row1 += 1


demand.cell(1, 1, 'From')
demand.cell(1, 2, 'To')
demand.cell(1, 3, 'Demand')

row2 = 2
for value1 in range(m * n):
    for value2 in range(n * m):
        demand.cell(row2, 1, value1+1)
        demand.cell(row2, 2, value2+1)
        demand.cell(row2, 3, net.get_demand(value1, value2))
        row2 += 1


wb.save(path)
